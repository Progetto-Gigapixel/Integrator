import json
import os

from openpyxl import Workbook, load_workbook

from core.utils.core_utils import delta_e_2000_colormath, get_colorchecker_24_patch_name
from locales.localization import _
from log.logger import logger
from utils.utils import convert_numpy
import numpy as np

from .model import ComparisonStep


class PatchComparer:
    def __init__(self, result_file, reference, image_path, process_cropped_image_only):
        """
        Compares the patches with the target reference.

        :param result_file: The file to save the comparison results
        :param reference: The reference object for the target reference
        """
        self.result_file = result_file
        self.reference = [reference[str(i)] for i in range(1, len(reference) + 1)]
        self.comparison_results = {"Corrections": {}}
        self.image_path = image_path
        self.process_cropped_image_only = process_cropped_image_only

    def compare_patches(self, patches):
        """
        Compares the patches with the target reference and calculates the ΔE00 values.

        :param patches: The patches to compare
        """
        delta_e_sum = 0
        delta_e_max = 0
        delta_e_min = 999999.0
        patches_lenght = len(patches[0])

        deltas_e_array = []
        for patch_number, (campione, patch_info) in enumerate(zip(self.reference, patches[0])):
            position = patch_info["cropped_image_position"] if self.process_cropped_image_only else patch_info["absolute_position"]
            size = patch_info["size"]

            corrected_rgb = patch_info["rgb_values"]
            reference_rgb = campione

            # reference_lab = rgb_2_lab(reference_rgb)
            # corrected_lab = rgb_2_lab(corrected_rgb)
            # delta_e = ciede2000(reference_lab, corrected_lab, kH=0.99, kC=0.99, kL=0.99)
            delta_e = delta_e_2000_colormath(reference_rgb, corrected_rgb)
            if patches_lenght== 24:
                patch_name = get_colorchecker_24_patch_name(patch_number)
            else:
                # TOD0 ::::
                patch_name = ""

            comparison_result = {
                "patch_number": patch_number + 1,
                "patch_name": patch_name,
                "position": position,
                "size": size,
                "correctedRGB": corrected_rgb,
                "referenceRGB": reference_rgb,
                "deltaE": delta_e,
            }
            # Append delta_e to do some stats afterwards
            deltas_e_array.append(delta_e)
            self.add_comparison_result(comparison_result)

        # Calculate the mean and max ΔE00 values
        delta_e_mean = np.mean(deltas_e_array)
        delta_e_standard_deviation = np.std(deltas_e_array)
        delta_e_max = np.max(deltas_e_array)
        delta_e_min = np.min(deltas_e_array)
        self.comparison_results["Corrections"][
            self.current_step
        ].delta_e_mean = delta_e_mean
        self.comparison_results["Corrections"][
            self.current_step
        ].delta_e_sd = delta_e_standard_deviation
        self.comparison_results["Corrections"][
            self.current_step
        ].delta_e_max = delta_e_max
        self.comparison_results["Corrections"][
            self.current_step
        ].delta_e_min = delta_e_min

        logger.debug(
            _("Mean DE00 for step {step}: {delta_e_mean:.4f}").format(
                step=self.current_step, delta_e_mean=delta_e_mean
            )
        )
        return delta_e_mean

    def add_comparison_result(self, comparison_result):
        """
        Adds a comparison result to the comparison results.

        :param comparison_result: The comparison result to add
        """
        self.comparison_results["Corrections"].setdefault(
            self.current_step, ComparisonStep()
        )
        self.comparison_results["Corrections"][self.current_step].add_patch(
            comparison_result
        )

    def set_extension_processed(self, file_extension):
        self.comparison_results["Corrections"]


    def set_step_applied(self, step_applied: bool):
        self.comparison_results["Corrections"].setdefault(
            self.current_step, ComparisonStep()
        )
        self.comparison_results["Corrections"][self.current_step].step_applied = step_applied


    def save_analytics_to_xls(self):
        # Load / Create xls
        xls_file_path = os.path.join(
            os.path.split(self.result_file)[0], "Analytics.xlsx"
        )
        image_filename = os.path.split(self.image_path)[1]

        if os.path.exists(xls_file_path):
            workbook = load_workbook(xls_file_path)
            sheet = workbook.active
            # Find the last row
            last_row = sheet.max_row
        else:
            workbook = Workbook()
            sheet = workbook.active
            # Write headers
            sheet.cell(1, 1, "Image")
            sheet.cell(1, 2, "Start-time")
            sheet.cell(1, 3, "End-time")
            for col, value in enumerate(
                self.comparison_results["Corrections"].items(), start=1
            ):
                de = value[1].delta_e_mean
                sheet.cell(row=1, column=col + 3, value=de)
            last_row = 1

        sheet.cell(row=last_row + 1, column=1, value=image_filename)
        # Write values in the last row of the file
        for col, values in enumerate(
            self.comparison_results["Corrections"].items(), start=1
        ):
            sheet.cell(row=last_row + 1, column=col + 3, value=values[1].delta_e_mean)

        # Salva le modifiche
        workbook.save(xls_file_path)
        workbook.close()

    def save_comparison_results(self):
        """
        Saves the comparison results to the result file.
        """
        results_to_save = {
            "Corrections": {
                step: comparison_step.to_dict()
                for step, comparison_step in self.comparison_results[
                    "Corrections"
                ].items()
            }
        }

        with open(self.result_file, "w") as file:
            json.dump(results_to_save, file, default=convert_numpy, indent=4)

        self.save_analytics_to_xls()

    def run(self, patches, current_step):
        """
        Runs the comparison process.

        :param patches: The patches to compare
        :param current_step: The current step
        """
        self.current_step = current_step

        return self.compare_patches(patches)
